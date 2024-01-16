from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import load_workbook

from .models import Bank, BalanceSheet
from .utils import (
    parse_excel_dates, create_or_update_balance_sheet, process_excel_row,
)

START_ROW = 9


# Представление для отображения данных по балансовому отчету
@csrf_exempt
@login_required
def show_data(request, id, template='osp/show_data.html'):
    context = {'balance_sheet': get_object_or_404(BalanceSheet, id=id)}
    return render(request, template, context)


# Представление для импорта данных из Excel-файла
@csrf_exempt
@login_required
def import_data(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']

        try:
            workbook = load_workbook(excel_file)
            sheet = workbook.active

            total_rows = sheet.max_row
            start_period, end_period = parse_excel_dates(sheet['A3'].value)

            # Импортируем банк из Excel-файла
            bank, _ = Bank.objects.get_or_create(name=sheet['A1'].value)
            # Импортируем ведомость из Excel-файла
            balance_sheet = create_or_update_balance_sheet(
                excel_file, bank, start_period, end_period
            )

            current_joint_code = None
            current_class = None
            # Проходимся циклом по строкам в таблице и парсим данные в СУБД
            for cell_values in sheet.iter_rows(
                min_row=START_ROW, max_row=total_rows, values_only=True
            ):
                current_class = process_excel_row(
                    cell_values, balance_sheet,
                    current_joint_code, current_class
                )

            workbook.close()
            messages.success(request, "Данные успешно импортированы.")
            return redirect('http://127.0.0.1:8000/admin/osv/balancesheet/')

        except Exception as e:
            messages.error(request, f"Ошибка чтения Excel-файла: {e}")
            return redirect('http://127.0.0.1:8000/admin/osv/balancesheet/')
